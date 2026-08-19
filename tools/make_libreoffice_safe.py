#!/usr/bin/env python3
"""Make the Third Parties Risk Evaluation Tool work in LibreOffice.

THE PROBLEM (verified, not theoretical)

The workbook computes the Supply Risk score with two formulas that only Excel
can evaluate:

    'Driver Configuration'!G2 = IFERROR(SUM(_xlfn._xlws.FILTER(D5:D17,
                                        C5:C17="YES","-")),"-") + D6
    'Driver Configuration'!C7 = 'Supply Risk Drivers'!H7:H17      (spilled)
    'Driver Configuration'!G3 = _xlfn.IFS(G2>=75%,"VERY CRITICAL", ...)

FILTER is a dynamic-array function; LibreOffice Calc has no dynamic arrays, so
on open it yields #VALUE! in G2 and, by cascade, in G3. The tool therefore
shows no score and no risk level in LibreOffice, whatever this project does.

THE FIX

Replace them with formulas that are identical in meaning and evaluate in both
Excel and LibreOffice:

    C7..C17 : one plain reference per row instead of one spilled range
    G2      : SUMPRODUCT((C5:C17="YES")*D5:D17) + D6
    G3      : nested IF instead of IFS

SUMPRODUCT is the standard pre-dynamic-array idiom for "sum where condition":
(C5:C17="YES") is an array of TRUE/FALSE, multiplying by D5:D17 coerces it to
1/0, and the sum is the same number FILTER+SUM produced. Nothing about the
model changes - only the dialect.

Usage:
    python tools/make_libreoffice_safe.py "Third Parties Risk Evaluation Tool v2.0.xlsx"
    python tools/make_libreoffice_safe.py in.xlsx -o out.xlsx

The original is never modified: the output is a new file with the suffix
`_lo` unless -o says otherwise.
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path

from openpyxl import load_workbook

SCORE_SHEET = "Driver Configuration"
ANSWER_SHEET = "Supply Risk Drivers"
FIRST_ROW, LAST_ROW = 5, 17

#: Answer column of 'Supply Risk Drivers'. H since tool v2.1, which inserted an
#: "ENISA 5G Security Controls Matrix" column before ANSWER; G before that. Read
#: from drivers.yaml when it can be found, so this tool cannot drift from the
#: config the rest of the project uses.
ANSWER_COL = "H"


def _answer_col() -> str:
    """Answer column from config/drivers.yaml, falling back to the constant."""
    cfg = Path(__file__).resolve().parents[1] / "config" / "drivers.yaml"
    try:
        import yaml
        data = yaml.safe_load(cfg.read_text(encoding="utf-8"))
        cell = data["drivers"][0]["answer_cell"]
        return "".join(c for c in cell if c.isalpha())
    except Exception:                                # noqa: BLE001 - best effort
        return ANSWER_COL


def convert(src: Path, dst: Path) -> dict[str, str]:
    shutil.copyfile(src, dst)
    wb = load_workbook(dst)
    ws = wb[SCORE_SHEET]
    changed: dict[str, str] = {}
    answer_col = _answer_col()

    # 1. the spilled range becomes one reference per row.
    #    Rows 5 and 6 already hold their own formulas and are left alone.
    for row in range(7, LAST_ROW + 1):
        formula = f"='{ANSWER_SHEET}'!{answer_col}{row}"
        ws.cell(row=row, column=3, value=formula)
        changed[f"C{row}"] = formula

    # 2. the weighted sum, without dynamic arrays
    g2 = (f'=SUMPRODUCT((C{FIRST_ROW}:C{LAST_ROW}="YES")*D{FIRST_ROW}:D{LAST_ROW})+D6')
    ws["G2"] = g2
    changed["G2"] = g2

    # 3. the risk level, without IFS
    g3 = ('=IF(G2>=0.75,"VERY CRITICAL",IF(G2>=0.5,"CRITICAL",'
          'IF(G2>=0.25,"SIGNIFICANT","NOT CRITICAL")))')
    ws["G3"] = g3
    changed["G3"] = g3

    wb.save(dst)
    return changed


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("tool", help="path to the Third Parties Risk Evaluation Tool .xlsx")
    p.add_argument("-o", "--output", help="output path (default: <name>_lo.xlsx)")
    args = p.parse_args()

    src = Path(args.tool)
    dst = Path(args.output) if args.output else src.with_name(src.stem + "_lo.xlsx")
    changed = convert(src, dst)

    print(f"{src.name} -> {dst}")
    for cell, formula in changed.items():
        print(f"  {cell:<5} {formula}")
    print("\nOpen the result in LibreOffice: G2 and G3 now evaluate.")
    print("The formulas remain valid in Excel too, so one file serves both.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
