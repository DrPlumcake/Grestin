"""Writing back into the tool is the riskiest part of the pipeline, because a
mistake damages the instrument the whole process depends on. These tests pin
the three invariants:

  1. only 'Supply Risk Drivers'!G5:G17 is written;
  2. the array formula in 'Driver Configuration'!G2 and the data validations
     survive the openpyxl round-trip;
  3. REVIEW never becomes an answer, and an already-filled cell is never
     overwritten.

Point to the real tool with TPRM_TOOL_XLSX=/path/to/tool.xlsx to run them;
otherwise they skip, so CI stays green without the (internal) workbook.
"""

import os
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

from grestin.config import Config
from grestin.hub.prefill import EVIDENCE_SHEET, prefill
from grestin.hub.scoring import projected_score, score
from grestin.models import Source, Verdict

TOOL = os.environ.get("TPRM_TOOL_XLSX", "")
pytestmark = pytest.mark.skipif(
    not TOOL or not Path(TOOL).exists(),
    reason="set TPRM_TOOL_XLSX to the Third Parties Risk Evaluation Tool to run these",
)


@pytest.fixture
def cfg():
    return Config.load()


@pytest.fixture
def summary(cfg):
    findings = [
        cfg.make_finding(type="kev_on_exposed_service", source=Source.KEV,
                         subject="CVE-2024-3400 on 203.0.113.7:443", evidence={"kev": True},
                         strength="strong"),
        cfg.make_finding(type="sanctions_match_fuzzy", source=Source.OPENSANCTIONS,
                         subject="Acme Holding Ltd", evidence={"score": 0.81},
                         strength="moderate"),
    ]
    return score(findings, cfg)


def test_default_mode_never_touches_the_answer_column(tmp_path, cfg, summary):
    """The conservative default: advice beside the question, no answer written."""
    out = tmp_path / "suggestions.xlsx"
    result = prefill(TOOL, out, summary, cfg, "Acme RAN S.p.A.", "TEST00",
                     projected_score(summary, cfg, {}))
    answers = load_workbook(out)[cfg.meta["answer_sheet"]]
    assert result["mode"] == "suggestion"
    assert result["cells_written"] == []
    assert answers["G10"].value in (None, "-")          # answer untouched
    assert answers["H10"].value == "SUGGEST_YES"        # advice present
    assert answers["I10"].value == "strong"
    assert "kev_on_exposed_service" in answers["J10"].value
    assert answers["H15"].value == "REVIEW"


def test_writes_only_suggest_yes_answers(tmp_path, cfg, summary):
    out = tmp_path / "prefilled.xlsx"
    result = prefill(TOOL, out, summary, cfg, "Acme RAN S.p.A.", "TEST01",
                     projected_score(summary, cfg, {}), write_answers=True)

    wb = load_workbook(out)
    answers = wb[cfg.meta["answer_sheet"]]
    # vuln_exposure_mgmt -> G10, SUGGEST_YES on a strong KEV signal
    assert answers["G10"].value == "YES"
    assert answers["G10"].comment is not None
    assert "SUGGEST_YES" in answers["G10"].comment.text
    # ownership_due_diligence -> G15, only REVIEW: advised in H, never answered in G
    assert answers["G15"].value in (None, "-")
    assert answers["H15"].value == "REVIEW"
    assert [w["driver"] for w in result["cells_written"]] == ["vuln_exposure_mgmt"]


def _formula(cell_value) -> str:
    return cell_value.text if isinstance(cell_value, ArrayFormula) else str(cell_value or "")


def test_scoring_model_and_validations_survive(tmp_path, cfg, summary):
    """The write-back must not damage the workbook's own model.

    Accepts either dialect, because tools/make_libreoffice_safe.py replaces the
    Excel-only dynamic array with a SUMPRODUCT of identical meaning: the test
    asserts that the score cell still sums the weighted YES answers and that
    the answer column still feeds it, not which function does the summing.
    """
    out = tmp_path / "prefilled.xlsx"
    prefill(TOOL, out, summary, cfg, "Acme", "TEST02", write_answers=True)
    wb = load_workbook(out)
    dc = wb[cfg.meta["score_sheet"]]

    g2 = _formula(dc[cfg.meta["score_cell"]].value)
    assert ("FILTER" in g2) or ("SUMPRODUCT" in g2), g2
    assert "D6" in g2                               # data classification still added
    assert cfg.meta["answer_sheet"] in _formula(dc["C7"].value)

    ranges = {str(dv.sqref) for dv in wb[cfg.meta["answer_sheet"]].data_validations.dataValidation}
    assert {"G5", "G6", "G7:G17"} <= ranges


def test_never_overwrites_an_existing_answer(tmp_path, cfg, summary):
    seeded = tmp_path / "seeded.xlsx"
    shutil.copyfile(TOOL, seeded)
    wb = load_workbook(seeded)
    wb[cfg.meta["answer_sheet"]]["G10"] = "NO"      # the compiler already decided
    wb.save(seeded)

    out = tmp_path / "prefilled.xlsx"
    result = prefill(seeded, out, summary, cfg, "Acme", "TEST03", write_answers=True)
    assert load_workbook(out)[cfg.meta["answer_sheet"]]["G10"].value == "NO"
    assert any("already filled" in f["reason"] for f in result["cells_flagged"])


def test_master_file_is_never_modified(tmp_path, cfg, summary):
    before = Path(TOOL).read_bytes()
    prefill(TOOL, tmp_path / "prefilled.xlsx", summary, cfg, "Acme", "TEST04")
    assert Path(TOOL).read_bytes() == before


def test_evidence_sheet_is_added_without_formulas(tmp_path, cfg, summary):
    out = tmp_path / "prefilled.xlsx"
    prefill(TOOL, out, summary, cfg, "Acme", "TEST05", projected_score(summary, cfg, {}))
    ws = load_workbook(out)[EVIDENCE_SHEET]
    values = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert not any(v.startswith("=") for v in values)   # nothing to recalculate
    assert any("PASSIVE OSINT" in v for v in values)


def test_non_boolean_driver_is_never_answered(tmp_path, cfg):
    """systems_access has an access-type answer list, not YES/NO: even a strong
    signal must not make the layer write into it."""
    findings = [cfg.make_finding(type="kev_on_exposed_service", source=Source.KEV,
                                 subject="CVE-1", evidence={}, strength="strong")]
    s = score(findings, cfg)
    for v in s.verdicts:
        if v.driver_id == "systems_access":
            assert v.verdict is Verdict.NOT_OBSERVABLE
    out = tmp_path / "p.xlsx"
    prefill(TOOL, out, s, cfg, "Acme", "TEST06", write_answers=True)
    assert load_workbook(out)[cfg.meta["answer_sheet"]]["G5"].value == "MAINTENANCE"
