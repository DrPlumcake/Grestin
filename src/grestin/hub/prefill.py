"""Write the CTI verdicts into a copy of the Third Parties Risk Evaluation Tool.

TWO MODES, and the default is the conservative one.

`suggestion` (default) - the layer does not touch the answer column at all. It
adds three columns beside it (H, I, J of 'Supply Risk Drivers') holding the
proposed verdict, the signal strength and the rationale, plus a `CTI Evidence`
sheet with the run summary. The person compiling the questionnaire reads the
suggestion next to the question and answers column G themselves. This mirrors
how the process actually works: Phase 1 is filled in by internal functions
(Procurement, the Requesting function, Security), not by the supplier, and the
signature on that cell must stay human.

`answers` (--write-answers) - additionally writes `YES` into column G for every
SUGGEST_YES verdict, with a cell comment carrying the evidence. Useful for the
demo, because it makes the score move on screen; not what you would deploy.

CONSTRAINTS DISCOVERED IN THE WORKBOOK

  1. The only input cells are 'Supply Risk Drivers'!G5:G17. Column C of
     'Driver Configuration' is derived, so writing there destroys the model.
  2. Each G cell carries a data-validation list, so a written value must belong
     to the driver's `answer_domain` in drivers.yaml. G5 and G6 are not YES/NO
     at all, and the layer never proposes a value for them.
  3. The original workbook computes the score with `_xlfn._xlws.FILTER`, which
     LibreOffice cannot evaluate (it has no dynamic arrays): opened there, G2
     and G3 read #VALUE!. That is a property of the tool, not of this code.
     Run `tools/make_libreoffice_safe.py` once to convert those formulas into
     SUMPRODUCT/IF equivalents that both suites evaluate. `prefill` detects the
     Excel-only formula and reports it so the CLI can warn.

In both modes the master file is copied first and never modified.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.formula import ArrayFormula

from ..config import Config
from ..models import Verdict
from .scoring import ScoreSummary

CTI_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
EVIDENCE_SHEET = "CTI Evidence"

SUGGESTION_COLS = {"verdict": 9, "strength": 10, "rationale": 11}   # I, J, K
HEADER_ROW = 4


def working_copy(template: str | Path, tool_dir: str | Path, slug: str,
                 config: Config, fresh: bool = False,
                 keep_template_answers: bool = False) -> Path:
    """Return the persistent per-supplier workbook, creating it if needed.

    One master template, one working copy per third party, no manual copying.
    The working copy is where the compiler actually answers, so the next run
    reads those answers back as `declared` and reports the delta against them
    - which is what makes the declared-vs-projected comparison mean anything.

    On creation the answer column is cleared, because a template usually
    carries whatever was last typed into it: without this, every supplier
    inherits the same phantom declared score. Pass
    `keep_template_answers=True` when the template legitimately holds defaults.
    """
    template, dest = Path(template), Path(tool_dir) / f"{slug}.xlsx"
    dest.parent.mkdir(parents=True, exist_ok=True)

    if dest.exists() and not fresh:
        return dest

    shutil.copyfile(template, dest)
    if keep_template_answers:
        return dest

    wb = load_workbook(dest)
    ws = wb[config.meta["answer_sheet"]]
    for driver in config.drivers.values():
        ws[driver.answer_cell] = "-"
    for col in SUGGESTION_COLS.values():             # drop stale CTI columns
        for row in range(HEADER_ROW, HEADER_ROW + 15):
            ws.cell(row=row, column=col).value = None
    if EVIDENCE_SHEET in wb.sheetnames:
        del wb[EVIDENCE_SHEET]
    wb.save(dest)
    return dest


def read_declared_answers(tool_path: str | Path, config: Config) -> dict[str, str]:
    """Read the answers already given in the tool: driver_id -> answer.

    Preferable to a hand-kept YAML, because the declared-vs-projected delta in
    the report is then computed against the actual state of the workbook the
    compiler is working in. Empty and "-" cells are omitted.
    """
    wb = load_workbook(tool_path, data_only=True)
    ws = wb[config.meta["answer_sheet"]]
    out: dict[str, str] = {}
    for driver in config.drivers.values():
        value = ws[driver.answer_cell].value
        if value in (None, "", "-"):
            continue
        out[driver.id] = str(value).strip()
    return out


def uses_excel_only_formulas(tool_path: str | Path, config: Config) -> bool:
    """True if the score cell still holds the dynamic-array version."""
    wb = load_workbook(tool_path)
    cell = wb[config.meta["score_sheet"]][config.meta["score_cell"]].value
    text = cell.text if isinstance(cell, ArrayFormula) else str(cell or "")
    return "FILTER" in text


def prefill(
    tool_path: str | Path,
    out_path: str | Path,
    summary: ScoreSummary,
    config: Config,
    target_name: str,
    run_id: str,
    projection: dict[str, Any] | None = None,
    write_answers: bool = False,
) -> dict[str, Any]:
    """Copy the tool, add the CTI suggestion columns and the evidence sheet."""
    tool_path, out_path = Path(tool_path), Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # Updating a workbook in place is legitimate for the per-third-party
    # working copy; copying is for every other case, so that a master template
    # is never touched.
    if tool_path.resolve() != out_path.resolve():
        shutil.copyfile(tool_path, out_path)

    excel_only = uses_excel_only_formulas(tool_path, config)
    wb = load_workbook(out_path)                  # keep formulas: no data_only
    answers = wb[config.meta["answer_sheet"]]

    _write_suggestion_header(answers)

    suggested: list[dict[str, Any]] = []
    answered: list[dict[str, Any]] = []
    flagged: list[dict[str, Any]] = []

    for v in summary.verdicts:
        driver = config.drivers[v.driver_id]
        row = int("".join(c for c in driver.answer_cell if c.isdigit()))
        if v.verdict is Verdict.NOT_OBSERVABLE and not v.findings:
            continue

        # --- always: the advisory columns, next to the question -----------
        cell = answers.cell(row=row, column=SUGGESTION_COLS["verdict"], value=v.verdict.value)
        cell.fill = CTI_FILL if v.verdict is Verdict.SUGGEST_YES else REVIEW_FILL
        answers.cell(row=row, column=SUGGESTION_COLS["strength"],
                     value=v.max_strength.value if v.max_strength else None)
        rationale = answers.cell(row=row, column=SUGGESTION_COLS["rationale"],
                                 value=_rationale_text(v, run_id))
        rationale.alignment = Alignment(wrap_text=True, vertical="top")
        suggested.append({"driver": v.driver_id, "row": row, "verdict": v.verdict.value})

        # --- optional: the answer itself ----------------------------------
        if not write_answers or v.verdict is not Verdict.SUGGEST_YES:
            continue
        answer_cell = answers[driver.answer_cell]
        if driver.answer_domain and "YES" not in driver.answer_domain:
            flagged.append({"driver": v.driver_id, "reason": "non-boolean answer domain"})
            continue
        if answer_cell.value not in (None, "-", ""):
            flagged.append({"driver": v.driver_id,
                            "reason": f"cell already filled: {answer_cell.value!r}"})
            continue
        answer_cell.value = "YES"
        answer_cell.fill = CTI_FILL
        answer_cell.font = Font(bold=True)
        answer_cell.comment = Comment(_rationale_text(v, run_id), "grestin",
                                      height=220, width=420)
        answered.append({"driver": v.driver_id, "cell": driver.answer_cell,
                         "weight": v.weight, "strength": v.max_strength.value})

    _write_evidence_sheet(wb, summary, config, target_name, run_id, projection, write_answers)
    wb.save(out_path)

    return {
        "output": str(out_path),
        "mode": "answers" if write_answers else "suggestion",
        "suggestions_written": suggested,
        "cells_written": answered,
        "cells_flagged": flagged,
        "excel_only_formulas": excel_only,
        "recalculation": (
            "the score cell still uses _xlfn._xlws.FILTER: it reads #VALUE! in LibreOffice. "
            "Run tools/make_libreoffice_safe.py on the master workbook once."
            if excel_only else
            "portable formulas: the score recalculates in both Excel and LibreOffice."
        ),
    }


def _write_suggestion_header(ws) -> None:
    bold = Font(bold=True, name="Arial")
    labels = {"verdict": "CTI SUGGESTION", "strength": "SIGNAL", "rationale": "CTI EVIDENCE"}
    widths = {"verdict": 18, "strength": 12, "rationale": 80}
    for key, col in SUGGESTION_COLS.items():
        cell = ws.cell(row=HEADER_ROW, column=col, value=labels[key])
        cell.font = bold
        cell.fill = HEADER_FILL
        ws.column_dimensions[cell.column_letter].width = widths[key]


def _rationale_text(verdict, run_id: str) -> str:
    lines = [f"{verdict.verdict.value} - {verdict.rationale}"]
    for f in verdict.findings[:6]:
        lines.append(f"- [{f.source.value}] {f.type} on {f.subject} ({f.signal_strength.value})")
    if verdict.corroborating_pillars:
        lines.append(f"corroborated across: {', '.join(verdict.corroborating_pillars)}")
    lines.append(f"advisory only - evidence: evidence/{run_id}/index.jsonl")
    return "\n".join(lines)


def _write_evidence_sheet(wb, summary: ScoreSummary, config: Config, target_name: str,
                          run_id: str, projection: dict[str, Any] | None,
                          write_answers: bool) -> None:
    """A plain sheet - no formulas - so the numbers survive without recalculation."""
    if EVIDENCE_SHEET in wb.sheetnames:
        del wb[EVIDENCE_SHEET]
    ws = wb.create_sheet(EVIDENCE_SHEET)
    bold = Font(bold=True, name="Arial")
    for col, width in zip("ABCDE", (36, 16, 10, 14, 72), strict=False):
        ws.column_dimensions[col].width = width

    ws["A1"] = "CTI LAYER - PASSIVE OSINT SUPPORT TO PHASE 1"
    ws["A1"].font = Font(bold=True, size=13, name="Arial")

    meta_rows = [
        ("Target", target_name),
        ("Run id", run_id),
        ("Mode", "answers written" if write_answers else "suggestions only"),
        ("Evidence index", f"evidence/{run_id}/index.jsonl"),
        ("Addressable weight (CTI-observable drivers)", summary.addressable_weight),
        ("Weight proposed as YES", round(summary.suggested_weight, 4)),
        ("Weight flagged for REVIEW", round(summary.review_weight, 4)),
    ]
    if projection:
        meta_rows += [
            ("Declared score / level",
             f"{projection['declared_score']:.0%} - {projection['declared_level']}"),
            ("Projected score / level (if all SUGGEST_YES accepted)",
             f"{projection['projected_score']:.0%} - {projection['projected_level']}"),
            ("Delta introduced by the CTI layer", f"{projection['delta']:.0%}"),
            ("Crosses the Phase 2 threshold", str(projection["crosses_phase2_threshold"])),
        ]
    row = 3
    for label, value in meta_rows:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    for col, header in enumerate(
            ["Driver", "Verdict", "Weight", "Max strength", "Rationale and signals"], start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = bold
        cell.fill = HEADER_FILL
    row += 1

    for v in summary.verdicts:
        if v.verdict is Verdict.NOT_OBSERVABLE and not v.findings:
            continue
        ws.cell(row=row, column=1, value=v.driver_name)
        ws.cell(row=row, column=2, value=v.verdict.value)
        ws.cell(row=row, column=3, value=v.weight)
        ws.cell(row=row, column=4, value=v.max_strength.value if v.max_strength else None)
        detail = v.rationale + "".join(
            f"\n- [{f.source.value}] {f.type} / {f.subject} ({f.signal_strength.value})"
            for f in v.findings[:8])
        ws.cell(row=row, column=5, value=detail).alignment = Alignment(wrap_text=True,
                                                                      vertical="top")
        ws.cell(row=row, column=2).fill = (
            CTI_FILL if v.verdict is Verdict.SUGGEST_YES else REVIEW_FILL)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Method").font = bold
    ws.cell(row=row, column=5, value=(
        "Passive OSINT only: no active scanning was performed against the supplier. "
        "The layer proposes YES or REVIEW and never NO: a driver without a signal is "
        "NOT_OBSERVABLE, not compliant. Every answer cell remains the compiler's "
        "responsibility."
    )).alignment = Alignment(wrap_text=True, vertical="top")
